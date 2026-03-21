from __future__ import annotations

from pathlib import Path

from homie_core.rag.parsers import ParsedDocument, TableData, TextBlock, register_parser


@register_parser("xlsx")
def parse_xlsx(path: Path) -> ParsedDocument:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ParsedDocument(source_path=str(path), metadata={"error": "openpyxl not installed"})
    wb = load_workbook(str(path), read_only=True, data_only=True)
    blocks = []
    tables = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append([str(cell) if cell is not None else "" for cell in row])
        if not rows_data:
            continue
        headers = rows_data[0]
        data_rows = rows_data[1:]
        tables.append(TableData(headers=headers, rows=data_rows, caption=sheet_name))
        text = f"Sheet: {sheet_name}\n" + "\n".join(", ".join(row) for row in rows_data)
        blocks.append(TextBlock(content=text, block_type="paragraph"))
    wb.close()
    return ParsedDocument(
        text_blocks=blocks,
        metadata={"format": "xlsx", "sheets": len(wb.sheetnames)},
        tables=tables,
        source_path=str(path),
    )
